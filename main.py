import os
import re
import time
import threading
from datetime import datetime

import kivy
kivy.require('2.0.0')
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.spinner import Spinner
from kivy.uix.popup import Popup
from kivy.clock import Clock
from kivy.utils import platform

# Serial communication
import serial
import serial.tools.list_ports

# Excel export
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

# OpenCV & PyZbar for camera scanning if available
try:
    import cv2
    from pyzbar.pyzbar import decode as decode_barcode
    CAMERA_AVAILABLE = True
except Exception:
    CAMERA_AVAILABLE = False


class SauterAndroidApp(App):
    def build(self):
        self.title = "Sauter FC Nadzor (2-12 Gnezd)"
        
        self.ser = None
        self.is_connected = False
        self.measuring_active = False
        self.has_pulled = False
        
        self.total_cavities = 8
        self.current_cavity = 1
        self.start_threshold = 3.00
        self.current_peak = 0.0
        
        self.results_peak = {}
        self.raw_curves = {}
        self.curr_t, self.curr_f = [], []
        self.start_time = 0

        # Main Tablet Layout (Landscape 2 Columns)
        main_layout = BoxLayout(orientation='horizontal', padding=10, spacing=10)

        # LEFT COLUMN
        left_box = BoxLayout(orientation='vertical', spacing=6, size_hint_x=0.56)

        # 1. Header (DN, Part, Operator, Cavity selector)
        hdr_grid = GridLayout(cols=3, spacing=4, size_hint_y=None, height=130)
        
        # Row 1: DN + Scan
        hdr_grid.add_widget(Label(text="* DN:", bold=True, size_hint_x=0.25, font_size=15))
        self.in_wo = TextInput(multiline=False, font_size=16, size_hint_x=0.55)
        hdr_grid.add_widget(self.in_wo)
        btn_scan_wo = Button(text="📷", size_hint_x=0.2, font_size=18, background_color=(0.3, 0.5, 0.8, 1))
        btn_scan_wo.bind(on_press=lambda x: self.scan_barcode(self.in_wo, "Delovni Nalog"))
        hdr_grid.add_widget(btn_scan_wo)

        # Row 2: Part + Scan
        hdr_grid.add_widget(Label(text="* Artikel:", bold=True, size_hint_x=0.25, font_size=15))
        self.in_part = TextInput(multiline=False, font_size=16, size_hint_x=0.55)
        hdr_grid.add_widget(self.in_part)
        btn_scan_part = Button(text="📷", size_hint_x=0.2, font_size=18, background_color=(0.3, 0.5, 0.8, 1))
        btn_scan_part.bind(on_press=lambda x: self.scan_barcode(self.in_part, "Artikel / Orodje"))
        hdr_grid.add_widget(btn_scan_part)

        # Row 3: Operator + Cavity count spinner
        hdr_grid.add_widget(Label(text="* Operater:", bold=True, size_hint_x=0.25, font_size=15))
        self.in_op = TextInput(multiline=False, font_size=16, size_hint_x=0.45)
        hdr_grid.add_widget(self.in_op)
        
        self.spinner_cavities = Spinner(
            text='8 Gnezd',
            values=('2 Gnezdi', '4 Gnezda', '6 Gnezd', '8 Gnezd', '10 Gnezd', '12 Gnezd'),
            size_hint_x=0.30,
            background_color=(0.8, 0.6, 0.2, 1),
            font_size=13
        )
        self.spinner_cavities.bind(text=self.on_cavity_count_change)
        hdr_grid.add_widget(self.spinner_cavities)

        left_box.add_widget(hdr_grid)

        # 2. Connection Bar
        conn_box = BoxLayout(orientation='horizontal', spacing=5, size_hint_y=None, height=42)
        self.btn_connect = Button(text="POVEŽI SAUTER (USB)", background_color=(0.2, 0.6, 0.9, 1), bold=True)
        self.btn_connect.bind(on_press=self.toggle_connection)
        conn_box.add_widget(self.btn_connect)

        self.lbl_status = Label(text="Odklopljeno", color=(1, 0.3, 0.3, 1), bold=True, size_hint_x=0.4)
        conn_box.add_widget(self.lbl_status)
        left_box.add_widget(conn_box)

        # 3. Peak Display
        display_box = BoxLayout(orientation='vertical', padding=4, spacing=2, size_hint_y=None, height=155)
        self.lbl_active_cavity = Label(text="PRIPRAVLJEN ZA: GNEZDO 1", font_size=19, bold=True, color=(0.3, 0.7, 1, 1))
        display_box.add_widget(self.lbl_active_cavity)

        self.lbl_peak_display = Label(text="0.00 N", font_size=52, bold=True, color=(1, 0.2, 0.2, 1))
        display_box.add_widget(self.lbl_peak_display)

        self.lbl_live_sub = Label(text="Trenutna sila v živo: 0.00 N", font_size=13, color=(0.8, 0.8, 0.8, 1))
        display_box.add_widget(self.lbl_live_sub)
        left_box.add_widget(display_box)

        # 4. Giant Action Button
        self.btn_step = Button(
            text="▶ ZAČNI MERITEV GNEZDA 1",
            font_size=19,
            bold=True,
            background_color=(0.2, 0.8, 0.3, 1),
            size_hint_y=None,
            height=65,
            disabled=True
        )
        self.btn_step.bind(on_press=self.trigger_measure_action)
        left_box.add_widget(self.btn_step)

        # 5. Aux Buttons
        aux_box = BoxLayout(orientation='horizontal', spacing=5, size_hint_y=None, height=42)
        btn_repeat = Button(text="↺ Ponovi gnezdo", on_press=self.repeat_current_cavity)
        btn_reset = Button(text="⚠ Ponastavi vsa", on_press=self.reset_all)
        aux_box.add_widget(btn_repeat)
        aux_box.add_widget(btn_reset)
        left_box.add_widget(aux_box)

        # Export Button
        btn_save = Button(text="💾 Ročno shrani v Excel", size_hint_y=None, height=40, on_press=lambda x: self.save_excel(manual=True))
        left_box.add_widget(btn_save)

        # RIGHT COLUMN: Dynamic table (2-12 slots)
        self.right_box = BoxLayout(orientation='vertical', spacing=3, size_hint_x=0.44)
        self.table_labels = {}
        
        self.rebuild_table(8)

        main_layout.add_widget(left_box)
        main_layout.add_widget(self.right_box)

        self.request_android_permissions()
        return main_layout

    def on_cavity_count_change(self, spinner, text):
        count = int(re.search(r'\d+', text).group(0))
        self.total_cavities = count
        self.rebuild_table(count)

    def rebuild_table(self, count):
        self.right_box.clear_widgets()
        self.table_labels.clear()
        self.results_peak = {i: None for i in range(1, count + 1)}
        self.raw_curves = {i: {"t": [], "f": []} for i in range(1, count + 1)}
        self.current_cavity = 1
        self.measuring_active = False
        self.has_pulled = False

        self.right_box.add_widget(Label(text=f"PREGLED PO GNEZDIH ({count})", font_size=17, bold=True, size_hint_y=None, height=32))
        
        row_height = 28 if count > 8 else 36
        for i in range(1, count + 1):
            row = BoxLayout(orientation='horizontal', spacing=5, size_hint_y=None, height=row_height)
            lbl_name = Label(text=f"Gnezdo {i}:", halign='left', size_hint_x=0.45, font_size=15)
            lbl_val = Label(text="---", bold=True, size_hint_x=0.55, font_size=16, color=(1, 1, 0.4, 1))
            row.add_widget(lbl_name)
            row.add_widget(lbl_val)
            self.table_labels[i] = lbl_val
            self.right_box.add_widget(row)

        if hasattr(self, 'lbl_active_cavity'):
            self.lbl_active_cavity.text = "PRIPRAVLJEN ZA: GNEZDO 1"
            self.lbl_peak_display.text = "0.00 N"
            if self.is_connected:
                self.btn_step.text = "▶ ZAČNI MERITEV GNEZDA 1"
                self.btn_step.background_color = (0.2, 0.8, 0.3, 1)
                self.btn_step.disabled = False

    def request_android_permissions(self):
        if platform == 'android':
            from android.permissions import request_permissions, Permission
            request_permissions([
                Permission.WRITE_EXTERNAL_STORAGE,
                Permission.READ_EXTERNAL_STORAGE,
                Permission.CAMERA
            ])

    def scan_barcode(self, target_input, field_name):
        if not CAMERA_AVAILABLE:
            self.show_popup("Kamera", "Kamera/PyZbar knjižnica ni na voljo na tej napravi.")
            return
        
        def worker():
            cap = cv2.VideoCapture(0)
            if not cap.isOpened():
                Clock.schedule_once(lambda dt: self.show_popup("Napaka", "Kamere ni bilo mogoče odpreti!"))
                return
            
            scanned_code = None
            qr_detector = cv2.QRCodeDetector()

            while True:
                ret, frame = cap.read()
                if not ret: break
                
                try:
                    barcodes = decode_barcode(frame)
                    if barcodes:
                        scanned_code = barcodes[0].data.decode('utf-8')
                        break
                except Exception:
                    pass

                if not scanned_code:
                    val, _, _ = qr_detector.detectAndDecode(frame)
                    if val:
                        scanned_code = val
                        break

                cv2.putText(frame, f"Skeniraj {field_name}", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                cv2.imshow("Kamera Skener", frame)
                if cv2.waitKey(1) & 0xFF in [ord('q'), ord('Q'), 27]:
                    break

            cap.release()
            cv2.destroyAllWindows()

            if scanned_code:
                Clock.schedule_once(lambda dt: self.apply_scanned(target_input, scanned_code))

        threading.Thread(target=worker, daemon=True).start()

    def apply_scanned(self, target_input, text):
        target_input.text = text.strip()

    def find_serial_port(self):
        ports = list(serial.tools.list_ports.comports())
        if ports:
            return ports[0].device
        for dev in ['/dev/ttyUSB0', '/dev/ttyUSB1', '/dev/ttyACM0', '/dev/ttyACM1']:
            if os.path.exists(dev):
                return dev
        return None

    def toggle_connection(self, instance):
        if self.is_connected:
            self.is_connected = False
            self.measuring_active = False
            if self.ser and self.ser.is_open:
                try: self.ser.close()
                except: pass
            self.btn_connect.text = "POVEŽI SAUTER (USB)"
            self.lbl_status.text = "Odklopljeno"
            self.lbl_status.color = (1, 0.3, 0.3, 1)
            self.btn_step.disabled = True
        else:
            port = self.find_serial_port()
            if not port:
                self.show_popup("Napaka", "Sauter ni zaznan preko USB kabla!
Preverite OTG priklop.")
                return
            try:
                self.ser = serial.Serial(
                    port=port, baudrate=38400, bytesize=serial.EIGHTBITS,
                    parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout=0.1
                )
                self.ser.dtr = True
                self.ser.rts = True
                time.sleep(0.2)

                self.is_connected = True
                self.btn_connect.text = "PREKINI"
                self.lbl_status.text = f"Povezano\n({os.path.basename(port)})"
                self.lbl_status.color = (0.3, 1, 0.3, 1)
                self.btn_step.disabled = False
                
                threading.Thread(target=self.polling_loop, daemon=True).start()
            except Exception as e:
                self.show_popup("Napaka pri povezavi", str(e))

    def polling_loop(self):
        while self.is_connected and self.ser and self.ser.is_open:
            try:
                self.ser.write(b'?')
                time.sleep(0.035)

                if self.ser.in_waiting > 0:
                    raw = self.ser.read(self.ser.in_waiting).decode('latin1', errors='ignore')
                    match = re.search(r'[-+]?\b\d+(?:[\.,]\d+)?\b', raw)
                    if match:
                        num = float(match.group(0).replace(',', '.'))
                        force_val = abs(num)
                        Clock.schedule_once(lambda dt, val=force_val: self.process_sample(val))
            except Exception:
                break

    def process_sample(self, val):
        self.lbl_live_sub.text = f"Trenutna sila v živo: {val:.2f} N"

        if not self.measuring_active:
            return

        if not self.has_pulled:
            if val >= self.start_threshold:
                self.has_pulled = True
                self.start_time = time.time()
                self.curr_t = [0.0]
                self.curr_f = [val]
                self.current_peak = val
                self.lbl_peak_display.text = f"{self.current_peak:.2f} N"
        else:
            t = round(time.time() - self.start_time, 2)
            self.curr_t.append(t)
            self.curr_f.append(val)

            if val > self.current_peak:
                self.current_peak = val
                self.lbl_peak_display.text = f"{self.current_peak:.2f} N"

            if len(self.curr_f) > 3 and self.current_peak >= self.start_threshold:
                padla_sila = (val < self.current_peak * 0.40) or (val < self.start_threshold)
                if padla_sila:
                    self.finish_current_cavity()

    def validate_header(self):
        if not self.in_wo.text.strip():
            self.show_popup("Manjka podatek", "Obvezno vnesite ali skenirajte DN!")
            return False
        if not self.in_part.text.strip():
            self.show_popup("Manjka podatek", "Obvezno vnesite ali skenirajte Artikel!")
            return False
        if not self.in_op.text.strip():
            self.show_popup("Manjka podatek", "Obvezno vnesite Operaterja!")
            return False
        return True

    def trigger_measure_action(self, instance=None):
        if not self.is_connected or not self.validate_header():
            return

        if not self.measuring_active:
            self.measuring_active = True
            self.has_pulled = False
            self.current_peak = 0.0
            self.curr_t, self.curr_f = [], []
            self.lbl_peak_display.text = "0.00 N"
            self.btn_step.text = f"⏹ ZAKLJUČI GNEZDO {self.current_cavity}"
            self.btn_step.background_color = (0.9, 0.2, 0.2, 1)
        else:
            self.finish_current_cavity()

    def finish_current_cavity(self):
        self.measuring_active = False
        self.has_pulled = False
        
        self.results_peak[self.current_cavity] = self.current_peak
        self.raw_curves[self.current_cavity] = {
            "t": list(self.curr_t),
            "f": list(self.curr_f)
        }

        self.table_labels[self.current_cavity].text = f"{self.current_peak:.2f} N"
        self.table_labels[self.current_cavity].color = (0.2, 1, 0.2, 1)

        if self.current_cavity < self.total_cavities:
            self.current_cavity += 1
            self.lbl_active_cavity.text = f"PRIPRAVLJEN ZA: GNEZDO {self.current_cavity}"
            self.btn_step.text = f"▶ ZAČNI MERITEV GNEZDA {self.current_cavity}"
            self.btn_step.background_color = (0.2, 0.8, 0.3, 1)
        else:
            self.lbl_active_cavity.text = f"VSIH {self.total_cavities} GNEZD IZMERJENIH!"
            self.btn_step.text = "✔ MERITEV ZAKLJUČENA"
            self.btn_step.background_color = (0.2, 0.5, 0.9, 1)
            self.btn_step.disabled = True
            self.save_excel(manual=False)

    def repeat_current_cavity(self, instance=None):
        self.measuring_active = False
        self.has_pulled = False
        self.current_peak = 0.0
        self.lbl_peak_display.text = "0.00 N"
        self.results_peak[self.current_cavity] = None
        self.raw_curves[self.current_cavity] = {"t": [], "f": []}
        self.table_labels[self.current_cavity].text = "---"
        self.table_labels[self.current_cavity].color = (1, 1, 0.4, 1)
        self.btn_step.text = f"▶ ZAČNI MERITEV GNEZDA {self.current_cavity}"
        self.btn_step.background_color = (0.2, 0.8, 0.3, 1)

    def reset_all(self, instance=None):
        self.rebuild_table(self.total_cavities)

    def save_excel(self, manual=False):
        wo = self.in_wo.text.strip() or "Brez_DN"
        part = self.in_part.text.strip() or "Neznano"
        op = self.in_op.text.strip() or "Operater"
        now_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        date_display = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

        if platform == 'android':
            base_dir = "/sdcard/Documents/Meritve_Excel"
            if not os.path.exists("/sdcard/Documents"):
                base_dir = "/sdcard/Download/Meritve_Excel"
        else:
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            base_dir = os.path.join(desktop, "Meritve_Excel")
            
        os.makedirs(base_dir, exist_ok=True)
        file_path = os.path.join(base_dir, f"DN_{wo}_{now_str}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Poročilo Meritev"
        ws.views.sheetView[0].showGridLines = True

        font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_normal = Font(name="Calibri", size=11)
        
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        fill_sub_header = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        fill_stat = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        thin_side = Side(style='thin', color='BFBFBF')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        ws["B2"] = "POROČILO MERITVE SILE ODPIRANJA"
        ws["B2"].font = font_title

        info_data = [
            ("Delovni Nalog (DN):", wo),
            ("Artikel / Orodje:", part),
            ("Operater:", op),
            ("Število Gnezd:", f"{self.total_cavities}"),
            ("Datum in Čas:", date_display)
        ]
        
        row_idx = 4
        for label, val in info_data:
            ws.cell(row=row_idx, column=2, value=label).font = font_bold
            ws.cell(row=row_idx, column=2).fill = fill_sub_header
            ws.cell(row=row_idx, column=2).border = border_all
            ws.cell(row=row_idx, column=2).alignment = align_left

            c_val = ws.cell(row=row_idx, column=3, value=val)
            c_val.font = font_normal
            c_val.border = border_all
            c_val.alignment = align_center
            row_idx += 1

        row_idx += 2
        header_row = row_idx
        ws.cell(row=header_row, column=2, value="Gnezdo").font = font_header
        ws.cell(row=header_row, column=2).fill = fill_header
        ws.cell(row=header_row, column=2).alignment = align_center
        ws.cell(row=header_row, column=2).border = border_all

        ws.cell(row=header_row, column=3, value="Max Sila (N)").font = font_header
        ws.cell(row=header_row, column=3).fill = fill_header
        ws.cell(row=header_row, column=3).alignment = align_center
        ws.cell(row=header_row, column=3).border = border_all

        measured_values = []
        for i in range(1, self.total_cavities + 1):
            row_idx += 1
            val = self.results_peak.get(i)
            
            c_lbl = ws.cell(row=row_idx, column=2, value=f"Gnezdo {i}")
            c_lbl.font = font_normal
            c_lbl.alignment = align_center
            c_lbl.border = border_all

            c_val = ws.cell(row=row_idx, column=3, value=val if val is not None else "---")
            c_val.font = font_bold if val is not None else font_normal
            c_val.alignment = align_right
            c_val.border = border_all
            if isinstance(val, (int, float)):
                c_val.number_format = "0.00"
                measured_values.append(val)

        if measured_values:
            row_idx += 1
            ws.cell(row=row_idx, column=2, value="POVPREČJE").font = font_bold
            ws.cell(row=row_idx, column=2).fill = fill_stat
            ws.cell(row=row_idx, column=2).border = border_all
            c_avg = ws.cell(row=row_idx, column=3, value=sum(measured_values) / len(measured_values))
            c_avg.font = font_bold
            c_avg.fill = fill_stat
            c_avg.number_format = "0.00"
            c_avg.alignment = align_right
            c_avg.border = border_all

            row_idx += 1
            ws.cell(row=row_idx, column=2, value="MINIMALNA SILA").font = font_bold
            ws.cell(row=row_idx, column=2).fill = fill_stat
            ws.cell(row=row_idx, column=2).border = border_all
            c_min = ws.cell(row=row_idx, column=3, value=min(measured_values))
            c_min.font = font_bold
            c_min.fill = fill_stat
            c_min.number_format = "0.00"
            c_min.alignment = align_right
            c_min.border = border_all

            row_idx += 1
            ws.cell(row=row_idx, column=2, value="MAKSIMALNA SILA").font = font_bold
            ws.cell(row=row_idx, column=2).fill = fill_stat
            ws.cell(row=row_idx, column=2).border = border_all
            c_max = ws.cell(row=row_idx, column=3, value=max(measured_values))
            c_max.font = font_bold
            c_max.fill = fill_stat
            c_max.number_format = "0.00"
            c_max.alignment = align_right
            c_max.border = border_all

        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 22

        # Krivulje
        ws2 = wb.create_sheet(title="Surove Krivulje")
        chart = LineChart()
        chart.title = f"Krivulje Odpiranja (1 - {self.total_cavities}) - DN {wo}"
        chart.y_axis.title = "Sila (N)"
        chart.x_axis.title = "Čas (s)"
        chart.width = 19
        chart.height = 12

        col_c = 1
        has_curve_data = False
        for i in range(1, self.total_cavities + 1):
            t_vals = self.raw_curves[i]["t"]
            f_vals = self.raw_curves[i]["f"]
            if t_vals and f_vals:
                has_curve_data = True
                ws2.cell(row=1, column=col_c, value=f"G{i}_Cas(s)").font = font_bold
                ws2.cell(row=1, column=col_c+1, value=f"G{i}").font = font_bold
                max_r = len(t_vals) + 1
                for r_idx, (t_val, f_val) in enumerate(zip(t_vals, f_vals), start=2):
                    ws2.cell(row=r_idx, column=col_c, value=t_val)
                    ws2.cell(row=r_idx, column=col_c+1, value=f_val)
                data_ref = Reference(ws2, min_col=col_c+1, min_row=1, max_row=max_r)
                chart.add_data(data_ref, titles_from_data=True)
                col_c += 3

        if has_curve_data:
            ws.add_chart(chart, "E4")

        wb.save(file_path)
        self.show_popup("Poročilo shranjeno", f"Excel je shranjen v:\n{file_path}")

    def show_popup(self, title, message):
        box = BoxLayout(orientation='vertical', padding=10, spacing=10)
        box.add_widget(Label(text=message, halign='center'))
        btn_close = Button(text="V REDU", size_hint_y=None, height=45)
        box.add_widget(btn_close)
        popup = Popup(title=title, content=box, size_hint=(0.8, 0.45))
        btn_close.bind(on_press=popup.dismiss)
        popup.open()

if __name__ == "__main__":
    SauterAndroidApp().run()
